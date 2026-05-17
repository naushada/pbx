#!/usr/bin/env python3
"""
generate-subscriber-template.py — produce docs/subscribers-template.xlsx
from a single source of truth (this script).

Why an XLSX template at all: operators fill the bulk-import sheet in
Excel/Sheets. Plain CSV can't carry visual signals; this XLSX
highlights mandatory columns (flat_number, name, email) in YELLOW
so the operator can't miss them. The CSV upload endpoint accepts the
same header row when the operator does Save As → CSV.

Re-run after touching the mandatory-set:
    python3 -m venv /tmp/xlsx_venv
    /tmp/xlsx_venv/bin/pip install openpyxl
    /tmp/xlsx_venv/bin/python3 scripts/generate-subscriber-template.py

The script is deterministic — re-running with the same inputs
produces a byte-identical .xlsx (modulo a timestamp openpyxl writes
into core.xml, which is acceptable churn in git).
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import os
import sys

# Single source of truth — keep in sync with the cloud's
# handle_subscriber_import_POST validator
# (modules/module/pbx/src/microservice_pbx.cpp). The cloud requires
# flat_number, name, email; phone + role are optional.
COLUMNS = [
    {"name": "flat_number", "mandatory": True,  "example": "A-101"},
    {"name": "name",        "mandatory": True,  "example": "Asha Resident"},
    {"name": "email",       "mandatory": True,  "example": "asha@example.com"},
    {"name": "phone",       "mandatory": False, "example": "+91 98765 43210"},
    {"name": "role",        "mandatory": False, "example": "resident"},
]

OUT_PATH = os.path.join(os.path.dirname(__file__), "..", "docs",
                         "subscribers-template.xlsx")
OUT_PATH = os.path.normpath(OUT_PATH)


def main() -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Subscribers"

    # ── Header row ──────────────────────────────────────────────────────
    yellow = PatternFill(start_color="FFF59D", end_color="FFF59D",
                          fill_type="solid")  # Material Yellow 200
    grey   = PatternFill(start_color="E0E0E0", end_color="E0E0E0",
                          fill_type="solid")
    bold   = Font(bold=True)
    centre = Alignment(horizontal="center", vertical="center")
    thin   = Side(border_style="thin", color="999999")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col["name"])
        cell.fill      = yellow if col["mandatory"] else grey
        cell.font      = bold
        cell.alignment = centre
        cell.border    = border

    # ── Example data row ────────────────────────────────────────────────
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col["example"])
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Column widths — auto-fit to example + 2 chars padding ──────────
    for col_idx, col in enumerate(COLUMNS, start=1):
        width = max(len(col["name"]), len(col["example"])) + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze the header so it stays visible on scroll ────────────────
    ws.freeze_panes = "A2"

    # ── Legend on the second sheet ──────────────────────────────────────
    legend = wb.create_sheet("Legend")
    legend["A1"] = "Column"
    legend["B1"] = "Mandatory?"
    legend["C1"] = "Notes"
    for cell in (legend["A1"], legend["B1"], legend["C1"]):
        cell.font = bold
        cell.fill = grey
    for row_idx, col in enumerate(COLUMNS, start=2):
        legend.cell(row=row_idx, column=1, value=col["name"]).fill = (
            yellow if col["mandatory"] else grey)
        legend.cell(row=row_idx, column=2,
                     value="YES" if col["mandatory"] else "no")
        legend.cell(row=row_idx, column=3, value={
            "flat_number": "Must match an existing flats doc when role=resident",
            "name":        "Display label shown in the directory and UI",
            "email":       "Used for portal login (unique within a society)",
            "phone":       "Optional contact number",
            "role":        "resident (default) | guard | admin",
        }.get(col["name"], ""))
    for letter, width in (("A", 16), ("B", 12), ("C", 60)):
        legend.column_dimensions[letter].width = width

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
